"""XLSX 电子表格处理技能

支持创建、读取、编辑、分析 Excel 电子表格，支持公式和财务建模。
"""

import json
import re
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from anyclaw.skills.base import Skill


# 财务模型配色常量
COLOR_BLUE = "0000FF"      # 输入值
COLOR_BLACK = "000000"     # 公式
COLOR_GREEN = "008000"     # 跨表链接
COLOR_RED = "FF0000"       # 外部链接
COLOR_YELLOW_BG = "FFFF00" # 关键假设背景


class XlsxSkill(Skill):
    """Excel spreadsheet operations: create, read, edit, analyze, recalc

    Supports:
    - Creating spreadsheets with formulas (not hardcoded values)
    - Reading and converting to markdown/JSON/CSV
    - Editing cells while preserving formulas
    - Data analysis with statistics
    - Formula recalculation via LibreOffice
    - Error detection (#REF!, #DIV/0!, #VALUE!)
    """

    def __init__(self):
        super().__init__()
        self.name = "xlsx"

    async def execute(
        self,
        action: str = "",
        path: str = "",
        content: str = "",
        output: str = "",
        sheet: str = "",
        output_format: str = "markdown",
        cell_range: str = "",
        financial: bool = False,
        **kwargs,
    ) -> str:
        """Execute XLSX operation

        Args:
            action: Operation to perform (create, read, edit, analyze, recalc, info, errors)
            path: Input file path
            content: Content for create/edit operations
            output: Output file path (optional)
            sheet: Sheet name (default: active sheet)
            output_format: Output format for read (markdown/json/csv)
            cell_range: Cell range for analysis (e.g., "A1:D100")
            financial: Enable financial model color coding
            **kwargs: Additional parameters

        Returns:
            Operation result
        """
        if not action:
            return self._help()

        if not path:
            return "Please provide a file path"

        try:
            action_handlers = {
                "create": self._create,
                "read": self._read,
                "edit": self._edit,
                "analyze": self._analyze,
                "recalc": self._recalc,
                "info": self._info,
                "errors": self._errors,
            }

            handler = action_handlers.get(action.lower())
            if not handler:
                return f"Unknown action: {action}. Supported: {', '.join(action_handlers.keys())}"

            return await handler(
                path=path,
                content=content,
                output=output,
                sheet=sheet,
                output_format=output_format,
                cell_range=cell_range,
                financial=financial,
                **kwargs,
            )

        except Exception as e:
            return f"Error: {str(e)}"

    def _help(self) -> str:
        """Return help message"""
        return """XLSX Skill - Excel Spreadsheet Operations

Actions:
  create  - Create a new spreadsheet with data and formulas
  read    - Read spreadsheet content (returns markdown table)
  edit    - Edit cells in existing spreadsheet
  analyze - Analyze data with statistics and insights
  recalc  - Recalculate all formulas using LibreOffice
  info    - Get spreadsheet metadata
  errors  - Check for formula errors

Parameters:
  path       - File path (required)
  content    - Data for create, cell updates for edit
  output     - Output file path (optional)
  sheet      - Sheet name (default: active sheet)
  output_format - Output format (markdown/json/csv)
  cell_range    - Cell range for analysis (e.g., "A1:D100")
  financial  - Enable financial model color coding

Examples:
  xlsx create /path/to/budget.xlsx --content '| A | B | C |
                                                |---|---|---|
                                                | 1 | 2 | =A1+B1 |'
  xlsx read /path/to/spreadsheet.xlsx
  xlsx edit /path/to/file.xlsx --content "A1:100|B1:=SUM(A1:A10)"
  xlsx analyze /path/to/data.xlsx
  xlsx recalc /path/to/spreadsheet.xlsx
  xlsx errors /path/to/spreadsheet.xlsx
"""

    async def _create(
        self,
        path: str,
        content: str = "",
        output: str = "",
        sheet: str = "",
        output_format: str = "markdown",
        cell_range: str = "",
        financial: bool = False,
        **kwargs,
    ) -> str:
        """Create a new spreadsheet with data and formulas"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return "Error: openpyxl not installed. Run: pip install openpyxl"

        path_obj = Path(path).expanduser().resolve()
        path_obj.parent.mkdir(parents=True, exist_ok=True)

        # Parse content
        data = self._parse_content(content)
        if not data:
            return "Error: No content provided. Use JSON array or markdown table format."

        # Create workbook
        wb = Workbook()
        ws = wb.active
        if sheet:
            ws.title = sheet

        # Define styles for financial models
        blue_font = Font(color=COLOR_BLUE) if financial else None
        yellow_fill = PatternFill(start_color=COLOR_YELLOW_BG, end_color=COLOR_YELLOW_BG, fill_type="solid") if financial else None

        # Fill data
        formula_count = 0
        for row_idx, row in enumerate(data, start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Determine if formula or value
                cell_str = str(cell_value) if cell_value is not None else ""
                if cell_str.startswith("="):
                    cell.value = cell_str
                    formula_count += 1
                    # Formulas use black font
                    if financial:
                        cell.font = Font(color=COLOR_BLACK)
                else:
                    # Try to convert to number
                    cell.value = self._parse_value(cell_str)
                    # Input values use blue font in financial models
                    if financial and cell.value is not None and not isinstance(cell.value, str):
                        cell.font = blue_font

        # Auto-adjust column widths
        for col_idx in range(1, len(data[0]) + 1 if data else 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Save
        wb.save(str(path_obj))

        result = f"✓ Created spreadsheet: {path_obj}\n"
        result += f"  Rows: {len(data)}, Columns: {len(data[0]) if data else 0}\n"
        result += f"  Formulas: {formula_count}"
        if financial:
            result += "\n  Financial model styling applied"

        return result

    def _parse_content(self, content: str) -> List[List[Any]]:
        """Parse content from JSON or markdown table format"""
        content = content.strip()

        # Try JSON first
        if content.startswith("["):
            try:
                data = json.loads(content)
                if isinstance(data, list) and all(isinstance(row, list) for row in data):
                    return data
            except json.JSONDecodeError:
                pass

        # Try markdown table
        if "|" in content:
            return self._parse_markdown_table(content)

        # Try CSV-like
        if "\n" in content and "," in content:
            return [
                [self._parse_value(cell.strip()) for cell in row.split(",")]
                for row in content.split("\n")
                if row.strip()
            ]

        return []

    def _parse_markdown_table(self, content: str) -> List[List[Any]]:
        """Parse markdown table to 2D array"""
        lines = content.strip().split("\n")
        data = []

        for line in lines:
            line = line.strip()
            if not line or line.startswith("|---") or set(line) <= {"|", "-", ":", " "}:
                continue

            # Extract cells
            cells = [c.strip() for c in line.split("|")]
            cells = [c for c in cells if c]  # Remove empty strings

            if cells:
                data.append([self._parse_value(c) for c in cells])

        return data

    def _parse_value(self, value: str) -> Any:
        """Parse string value to appropriate type"""
        if not value or value == "-":
            return "" if value == "" else 0

        # Formula
        if value.startswith("="):
            return value

        # Boolean
        if value.lower() in ("true", "false"):
            return value.lower() == "true"

        # Number
        try:
            if "." in value:
                return float(value)
            return int(value)
        except ValueError:
            pass

        return value

    async def _read(
        self,
        path: str,
        content: str = "",
        output: str = "",
        sheet: str = "",
        output_format: str = "markdown",
        cell_range: str = "",
        financial: bool = False,
        **kwargs,
    ) -> str:
        """Read spreadsheet content"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl not installed. Run: pip install openpyxl"

        path_obj = Path(path).expanduser().resolve()

        if not path_obj.exists():
            return f"File not found: {path}"

        if path_obj.suffix.lower() not in (".xlsx", ".xlsm"):
            return f"Not an XLSX file: {path}"

        try:
            wb = load_workbook(str(path_obj), data_only=False)
            ws = wb[sheet] if sheet else wb.active

            # Determine range
            if cell_range:
                cells = ws[cell_range]
            else:
                cells = ws.iter_rows(values_only=False)

            # Extract data
            data = []
            formula_cells = []

            for row in cells:
                row_data = []
                for cell in row:
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        if cell_value.startswith("="):
                            row_data.append(cell_value)
                            formula_cells.append(cell.coordinate)
                        else:
                            row_data.append(cell.value)
                    else:
                        row_data.append("")
                data.append(row_data)

            # Format output
            if output_format == "json":
                result = json.dumps(data, indent=2, ensure_ascii=False)
            elif output_format == "csv":
                result = self._to_csv(data)
            else:
                result = self._to_markdown_table(data)

            # Add formula info
            if formula_cells:
                formula_info = f"Sheet: {ws.title}\nFormulas in: {', '.join(formula_cells[:10])}"
                if len(formula_cells) > 10:
                    formula_info += f" ... and {len(formula_cells) - 10} more"
                result = formula_info + "\n\n" + result

            # Limit output
            if len(result) > 50000:
                result = result[:50000] + "\n\n... [content truncated]"

            return f"Content of {path_obj.name}:\n\n{result}"

        except Exception as e:
            return f"Error reading spreadsheet: {str(e)}"

    def _to_markdown_table(self, data: List[List[Any]]) -> str:
        """Convert 2D array to markdown table"""
        if not data:
            return ""

        lines = []
        col_count = max(len(row) for row in data)

        for i, row in enumerate(data):
            # Pad row to max columns
            padded = row + [""] * (col_count - len(row))
            lines.append("| " + " | ".join(str(c) if c is not None else "" for c in padded) + " |")

            # Add separator after header
            if i == 0:
                lines.append("| " + " | ".join(["---"] * col_count) + " |")

        return "\n".join(lines)

    def _to_csv(self, data: List[List[Any]]) -> str:
        """Convert 2D array to CSV"""
        lines = []
        for row in data:
            csv_row = []
            for cell in row:
                cell_str = str(cell) if cell is not None else ""
                # Escape quotes and wrap in quotes if needed
                if "," in cell_str or '"' in cell_str or "\n" in cell_str:
                    cell_str = '"' + cell_str.replace('"', '""') + '"'
                csv_row.append(cell_str)
            lines.append(",".join(csv_row))
        return "\n".join(lines)

    async def _edit(
        self,
        path: str,
        content: str = "",
        output: str = "",
        sheet: str = "",
        output_format: str = "markdown",
        cell_range: str = "",
        financial: bool = False,
        **kwargs,
    ) -> str:
        """Edit cells in existing spreadsheet

        Content format: "A1:value|B2:=SUM(A1:A10)|C3:text"
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font
        except ImportError:
            return "Error: openpyxl not installed. Run: pip install openpyxl"

        path_obj = Path(path).expanduser().resolve()

        if not path_obj.exists():
            return f"File not found: {path}"

        if not content:
            return "Please provide cell updates in format: 'A1:value|B2:=SUM(A1:A10)'"

        try:
            wb = load_workbook(str(path_obj))
            ws = wb[sheet] if sheet else wb.active

            # Parse updates
            updates = self._parse_cell_updates(content)
            if not updates:
                return "No valid cell updates found. Use format: 'A1:value|B2:=formula'"

            # Apply updates
            for cell_ref, value in updates:
                cell = ws[cell_ref]

                # Check if formula
                if isinstance(value, str) and value.startswith("="):
                    cell.value = value
                    if financial:
                        cell.font = Font(color=COLOR_BLACK)
                else:
                    cell.value = self._parse_value(value) if isinstance(value, str) else value
                    if financial and isinstance(cell.value, (int, float)):
                        cell.font = Font(color=COLOR_BLUE)

            # Save
            output_path = Path(output).expanduser().resolve() if output else path_obj
            wb.save(str(output_path))

            return f"✓ Edited spreadsheet: {output_path}\n  Cells updated: {len(updates)}"

        except Exception as e:
            return f"Error editing spreadsheet: {str(e)}"

    def _parse_cell_updates(self, content: str) -> List[Tuple[str, str]]:
        """Parse cell updates from content string"""
        updates = []
        cell_pattern = re.compile(r"^([A-Z]+\d+):(.+)$", re.IGNORECASE)

        for part in content.split("|"):
            part = part.strip()
            match = cell_pattern.match(part)
            if match:
                cell_ref = match.group(1).upper()
                value = match.group(2).strip()
                updates.append((cell_ref, value))

        return updates

    async def _analyze(
        self,
        path: str,
        content: str = "",
        output: str = "",
        sheet: str = "",
        output_format: str = "markdown",
        cell_range: str = "",
        financial: bool = False,
        **kwargs,
    ) -> str:
        """Analyze spreadsheet data with statistics"""
        try:
            import pandas as pd
        except ImportError:
            return "Error: pandas not installed. Run: pip install pandas"

        path_obj = Path(path).expanduser().resolve()

        if not path_obj.exists():
            return f"File not found: {path}"

        try:
            # Read spreadsheet
            df = pd.read_excel(
                str(path_obj),
                sheet_name=sheet if sheet else 0,
                header=0 if not cell_range else None
            )

            if cell_range:
                # Parse range like "A1:D100"
                df = pd.read_excel(str(path_obj), sheet_name=sheet if sheet else 0, header=None)
                # TODO: Apply range filter

            result_lines = [f"Analysis of {path_obj.name}"]
            result_lines.append(f"Sheet: {sheet if sheet else 'default'}")
            result_lines.append(f"Shape: {df.shape[0]} rows × {df.shape[1]} columns")
            result_lines.append("")

            # Column info
            result_lines.append("Columns:")
            for col in df.columns:
                dtype = df[col].dtype
                non_null = df[col].count()
                result_lines.append(f"  - {col}: {dtype} ({non_null} non-null)")
            result_lines.append("")

            # Numeric statistics
            numeric_cols = df.select_dtypes(include=["number"]).columns
            if len(numeric_cols) > 0:
                result_lines.append("Numeric Statistics:")
                stats = df[numeric_cols].describe()
                result_lines.append(stats.to_string())
                result_lines.append("")

            # Sample data
            result_lines.append("First 5 rows:")
            result_lines.append(df.head().to_string())

            result = "\n".join(result_lines)

            # Limit output
            if len(result) > 50000:
                result = result[:50000] + "\n\n... [content truncated]"

            return result

        except Exception as e:
            return f"Error analyzing spreadsheet: {str(e)}"

    async def _recalc(
        self,
        path: str,
        content: str = "",
        output: str = "",
        sheet: str = "",
        output_format: str = "markdown",
        cell_range: str = "",
        financial: bool = False,
        **kwargs,
    ) -> str:
        """Recalculate all formulas using LibreOffice"""
        path_obj = Path(path).expanduser().resolve()

        if not path_obj.exists():
            return f"File not found: {path}"

        # Check for LibreOffice
        libreoffice_paths = [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            "libreoffice",
            "soffice",
        ]

        soffice = None
        for p in libreoffice_paths:
            if shutil.which(p):
                soffice = p
                break
            if Path(p).exists():
                soffice = p
                break

        if not soffice:
            return "Error: LibreOffice not found. Install: brew install --cask libreoffice"

        try:
            # Create temp directory for output
            with tempfile.TemporaryDirectory() as temp_dir:
                # Run LibreOffice in headless mode to recalculate
                cmd = [
                    soffice,
                    "--headless",
                    "--calc",
                    "--convert-to", "xlsx",
                    "--outdir", temp_dir,
                    str(path_obj),
                ]

                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=60,
                )

                if result.returncode != 0:
                    return f"LibreOffice error: {result.stderr}"

                # Copy result to output or overwrite original
                output_name = path_obj.stem + ".xlsx"
                temp_output = Path(temp_dir) / output_name

                output_path = Path(output).expanduser().resolve() if output else path_obj

                if temp_output.exists():
                    import shutil as shutil_module
                    shutil_module.copy2(temp_output, output_path)
                    return f"✓ Recalculated formulas: {output_path}"
                else:
                    return "Error: Recalculation failed - no output file generated"

        except subprocess.TimeoutExpired:
            return "Error: LibreOffice recalculation timed out"
        except Exception as e:
            return f"Error recalculating formulas: {str(e)}"

    async def _info(
        self,
        path: str,
        content: str = "",
        output: str = "",
        sheet: str = "",
        output_format: str = "markdown",
        cell_range: str = "",
        financial: bool = False,
        **kwargs,
    ) -> str:
        """Get spreadsheet metadata"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl not installed. Run: pip install openpyxl"

        path_obj = Path(path).expanduser().resolve()

        if not path_obj.exists():
            return f"File not found: {path}"

        try:
            wb = load_workbook(str(path_obj), data_only=False)

            info_lines = [f"Spreadsheet: {path_obj.name}"]
            info_lines.append(f"Path: {path_obj}")
            info_lines.append(f"Size: {path_obj.stat().st_size} bytes")
            info_lines.append("")

            # Workbook properties
            props = wb.properties
            if props.title:
                info_lines.append(f"Title: {props.title}")
            if props.creator:
                info_lines.append(f"Creator: {props.creator}")
            if props.created:
                info_lines.append(f"Created: {props.created}")
            if props.modified:
                info_lines.append(f"Modified: {props.modified}")
            if props.lastModifiedBy:
                info_lines.append(f"Last Modified By: {props.lastModifiedBy}")

            # Sheet info
            info_lines.append(f"\nSheets ({len(wb.sheetnames)}):")
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                formula_count = sum(
                    1 for row in ws.iter_rows() for cell in row
                    if cell.value and str(cell.value).startswith("=")
                )
                info_lines.append(f"  - {sheet_name}: {ws.max_row} rows × {ws.max_column} cols, {formula_count} formulas")

            return "\n".join(info_lines)

        except Exception as e:
            return f"Error getting spreadsheet info: {str(e)}"

    async def _errors(
        self,
        path: str,
        content: str = "",
        output: str = "",
        sheet: str = "",
        output_format: str = "markdown",
        cell_range: str = "",
        financial: bool = False,
        **kwargs,
    ) -> str:
        """Check for formula errors in spreadsheet"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: openpyxl not installed. Run: pip install openpyxl"

        path_obj = Path(path).expanduser().resolve()

        if not path_obj.exists():
            return f"File not found: {path}"

        # Common Excel error values
        error_values = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!", "#SPILL!", "#CALC!"}

        try:
            # Load with calculated values
            wb = load_workbook(str(path_obj), data_only=True)
            wb_formulas = load_workbook(str(path_obj), data_only=False)

            all_errors = []

            for sheet_name in wb.sheetnames:
                if sheet and sheet_name.lower() != sheet.lower():
                    continue

                ws = wb[sheet_name]
                ws_formulas = wb_formulas[sheet_name]

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell_value = str(cell.value).upper()
                            for error in error_values:
                                if error in cell_value:
                                    # Get the formula
                                    formula_cell = ws_formulas[cell.coordinate]
                                    formula = str(formula_cell.value) if formula_cell.value else ""

                                    all_errors.append({
                                        "sheet": sheet_name,
                                        "cell": cell.coordinate,
                                        "error": error,
                                        "formula": formula if formula.startswith("=") else "(value)",
                                    })

            if not all_errors:
                return f"✓ No formula errors found in {path_obj.name}"

            result_lines = [f"Found {len(all_errors)} error(s) in {path_obj.name}:"]
            for err in all_errors:
                result_lines.append(f"  [{err['sheet']}] {err['cell']}: {err['error']}")
                if err['formula'] != "(value)":
                    result_lines.append(f"    Formula: {err['formula']}")

            return "\n".join(result_lines)

        except Exception as e:
            return f"Error checking for errors: {str(e)}"
